import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import time
import re
import random
from datetime import datetime
from urllib.parse import quote

try:
    import cloudscraper
    USE_CLOUDSCRAPER = True
except ImportError:
    USE_CLOUDSCRAPER = False


class DigikeySpider:
    def __init__(self):
        # 优先使用 cloudscraper（自动绕过 Cloudflare）
        _use_cloudscraper = USE_CLOUDSCRAPER
        _init_success = False
        
        if _use_cloudscraper:
            try:
                self.session = cloudscraper.create_scraper(
                    browser={
                        'browser': 'chrome',
                        'platform': 'windows',
                        'mobile': False
                    }
                )
                print("[*] 使用 cloudscraper 模式（自动绕过 Cloudflare）")
                _init_success = True
            except Exception as e:
                print(f"[!] cloudscraper 初始化失败: {e}，回退到 requests")
        
        if not _init_success:
            self.session = requests.Session()
            self._setup_basic_headers()
            if not _use_cloudscraper:
                print("[!] 未安装 cloudscraper，使用 requests（可能被拦截）")
        
        self.timeout = 30
        
        # User-Agent 池，模拟不同浏览器
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
        ]
        
        # 随机延时范围（秒）
        self.min_delay = 1.5
        self.max_delay = 3.5

    def random_delay(self):
        """随机延时，模拟人类操作"""
        delay = random.uniform(self.min_delay, self.max_delay)
        time.sleep(delay)

    def _setup_basic_headers(self):
        """设置基础请求头"""
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'Accept-Encoding': 'gzip, deflate, br',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'Cache-Control': 'max-age=0',
            'DNT': '1',
            'Connection': 'keep-alive',
        })

    def get_random_headers(self):
        """获取随机请求头"""
        return {
            'User-Agent': random.choice(self.user_agents),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }

    def load_keywords(self, filepath):
        """从配置文件读取keywords"""
        with open(filepath, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]

    def search_product(self, keyword, retry=3):
        """搜索产品并获取详情页链接
        返回: (detail_url, current_page_stock, is_direct_detail)
        - detail_url: 详情页URL
        - current_page_stock: 如果直接跳转到详情页，当前页的库存
        - is_direct_detail: 是否直接跳转到详情页
        """
        search_url = f'https://www.digikey.cn/zh/products/result?keywords={keyword}'
        
        for attempt in range(retry):
            try:
                # 不自动跟随重定向，手动跟踪
                response = self.session.get(search_url, allow_redirects=False, timeout=self.timeout)
                
                # 检查是否被 Cloudflare 拦截
                if response.status_code == 403 or 'cf-' in response.headers.get('Server', '').lower():
                    if attempt < retry - 1:
                        print(f"  -> 检测到Cloudflare拦截，{5*(attempt+1)}秒后重试...")
                        time.sleep(5 * (attempt + 1))
                        continue
                    else:
                        print(f"  -> Cloudflare拦截次数过多，跳过")
                        return (None, None, False)

                # 跟踪重定向链
                current_url = search_url
                max_redirects = 10
                redirect_count = 0
                category_page_html = None  # 保存category页面用于查找a标签

                while redirect_count < max_redirects:
                    if response.status_code in (301, 302, 303, 307, 308):
                        # 获取重定向Location
                        location = response.headers.get('Location', '')
                        if not location:
                            break

                        # 处理相对路径
                        if location.startswith('/'):
                            location = 'https://www.digikey.cn' + location
                        elif not location.startswith('http'):
                            from urllib.parse import urljoin
                            location = urljoin(current_url, location)

                        current_url = location
                        redirect_count += 1

                        # 请求重定向目标页面
                        response = self.session.get(current_url, allow_redirects=False, timeout=self.timeout)

                        # 检查当前请求的URL类型（这是重定向后实际访问的URL）
                        if '/zh/products/detail/' in current_url:
                            # 直接跳转到详情页，获取当前响应的库存
                            stock = self._extract_stock_from_html(response.text, keyword)
                            return (current_url, stock, True)
                        elif '/zh/products/category/' in current_url:
                            # 跳转到分类页，保存当前页面内容，稍后查找a标签
                            category_page_html = response.text
                            continue
                        # 否则继续循环（可能还有更多重定向）
                    else:
                        break

                # 如果之前有跳转到category页面，在该页面查找a标签
                if category_page_html:
                    soup = BeautifulSoup(category_page_html, 'lxml')
                    link = soup.select_one('a.tss-css-1abf7dr-Link-anchor-buttonAnchor')
                    if link and link.get('href'):
                        detail_url = 'https://www.digikey.cn' + link['href']
                        return (detail_url, None, False)

                # 在当前最终页面查找链接
                soup = BeautifulSoup(response.text, 'lxml')
                link = soup.select_one('a.tss-css-1abf7dr-Link-anchor-buttonAnchor')
                if link and link.get('href'):
                    detail_url = 'https://www.digikey.cn' + link['href']
                    return (detail_url, None, False)

                return (None, None, False)
                
            except Exception as e:
                if attempt < retry - 1:
                    print(f"  -> 请求失败，{3*(attempt+1)}秒后重试...")
                    time.sleep(3 * (attempt + 1))
                    continue
                print(f"搜索 {keyword} 失败: {e}")
                return (None, None, False)

    def _extract_stock_from_html(self, html, keyword):
        """从HTML中提取库存数量"""
        soup = BeautifulSoup(html, 'lxml')

        # 查找现货数量，支持多种格式:
        # - "现货: 1,324" (数字在冒号后)
        # - "0 现货" (数字在前)
        # - "现货: 0" (数字在冒号后为0)
        
        # 方案1：匹配 "现货: 1,324" 或 "现货: 0"
        stock_span = soup.find('span', string=re.compile(r'现货'))
        if stock_span:
            # 先尝试 "数字 现货" 格式
            match = re.search(r'([\d,]+)\s+现货', stock_span.string)
            if match:
                return match.group(1).replace(',', '')
            
            # 再尝试 "现货: 数字" 格式
            match = re.search(r'现货[：:]\s*([\d,]+)', stock_span.string)
            if match:
                return match.group(1).replace(',', '')

        # 备选方案：查找所有span中的数字
        all_spans = soup.find_all('span')
        for span in all_spans:
            if span.string and '现货' in span.string:
                # 先尝试 "数字 现货" 格式
                match = re.search(r'([\d,]+)\s+现货', span.string)
                if match:
                    return match.group(1).replace(',', '')
                
                # 再尝试 "现货: 数字" 格式
                match = re.search(r'现货[：:]\s*([\d,]+)', span.string)
                if match:
                    return match.group(1).replace(',', '')

        return '未找到'

    def get_stock(self, detail_url, keyword, retry=3):
        """获取详情页的现货数量"""
        for attempt in range(retry):
            try:
                response = self.session.get(detail_url, timeout=self.timeout)
                
                # 检查是否被拦截
                if response.status_code == 403:
                    if attempt < retry - 1:
                        print(f"  -> 请求被拦截，{3*(attempt+1)}秒后重试...")
                        time.sleep(3 * (attempt + 1))
                        continue
                    else:
                        return '请求被拦截'
                
                return self._extract_stock_from_html(response.text, keyword)
            except Exception as e:
                if attempt < retry - 1:
                    print(f"  -> 获取失败，{3*(attempt+1)}秒后重试...")
                    time.sleep(3 * (attempt + 1))
                    continue
                print(f"获取 {keyword} 库存失败: {e}")
                return '获取失败'

    def crawl(self, keywords, progress_callback=None):
        """爬取所有关键词的产品信息
        
        Args:
            keywords: 关键词列表
            progress_callback: 进度回调函数，签名为 callback(current, total, keyword, result)
        """
        results = []

        for i, keyword in enumerate(keywords):
            print(f"[{i+1}/{len(keywords)}] 正在爬取: {keyword}")

            # 获取详情页URL和库存信息
            detail_url, direct_stock, is_direct = self.search_product(keyword)

            if is_direct:
                # 直接跳转到详情页，使用当前页面的库存
                stock = direct_stock
                print(f"  -> 直接跳转详情页，现货数量: {stock}")
            elif detail_url:
                # 需要访问详情页获取库存
                stock = self.get_stock(detail_url, keyword)
                print(f"  -> 现货数量: {stock}")
            else:
                stock = '未找到产品'
                print(f"  -> 未找到产品")

            result = {
                '规格': keyword,
                '现货数量': stock
            }
            results.append(result)
            
            # 进度回调
            if progress_callback:
                progress_callback(i + 1, len(keywords), keyword, result)

            # 随机延时（1.5-3.5秒），避免请求过快
            self.random_delay()

        return results

    def export_to_excel(self, data, output_path):
        """导出数据到Excel"""
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine='openpyxl')

        # 美化Excel
        wb = load_workbook(output_path)
        ws = wb.active

        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置数据区域样式和对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        wb.save(output_path)
        print(f"Excel已导出: {output_path}")


def main():
    spider = DigikeySpider()

    # 读取关键词
    keywords = spider.load_keywords('config/keywords.txt')
    print(f"已加载 {len(keywords)} 个关键词\n")

    # 爬取数据
    results = spider.crawl(keywords)

    # 导出Excel（文件名带时间戳）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'output/digikey_stock_{timestamp}.xlsx'
    spider.export_to_excel(results, output_file)


if __name__ == '__main__':
    main()
