"""
Digikey 爬虫 V2 — 基于 Playwright + 人工绕过 Cloudflare 验证
当检测到 CF 验证页面时，弹出对话框提示真人在浏览器中手动完成验证，
确认后程序继续执行后续流程。
"""
import time
import re
import random
import threading
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from urllib.parse import quote

from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ---------------------------------------------------------------------------
# Cloudflare 检测 & 人工干预
# ---------------------------------------------------------------------------

CF_CHECK_INTERVAL = 2          # 检测间隔（秒）
CF_TIMEOUT = 300               # 最长等待真人验证时间（秒）

# 标记真人是否已确认通过 CF 验证
_cf_passed_event = threading.Event()


def _show_cf_popup():
    """在独立线程中弹出 Tkinter 对话框，等待真人确认"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    root.attributes('-topmost', True)
    messagebox.showinfo(
        "Cloudflare 验证",
        "检测到 Cloudflare 人机验证！\n\n"
        "请在打开的浏览器中手动完成验证（勾选复选框等），\n"
        "验证通过后点击「确定」继续爬取。"
    )
    root.destroy()
    _cf_passed_event.set()


def _is_cf_challenge(page) -> bool:
    """判断当前页面是否是 Cloudflare 验证页面（只用轻量方法，避免 content()）"""
    try:
        title = page.title()
        if 'Just a moment' in title:
            return True
        if 'Attention Required' in title:
            return True
        # 通过 DOM 元素判断
        if page.locator('#challenge-running, #challenge-form, div.cf-browser-verification').count() > 0:
            return True
    except Exception:
        return False
    return False


def _is_digikey_normal(page) -> bool:
    """判断页面是否已经正常加载出 Digikey 内容（CF 验证已通过）"""
    try:
        # 详情页特征
        if page.locator('[data-testid="in-stock-title"]').count() > 0:
            return True
        # 搜索列表页特征
        if page.locator('table tbody tr').count() > 0:
            return True
        if page.locator('a[href*="/detail/"]').count() > 0:
            return True
        # URL 已经是 digikey 正常页面
        url = page.url
        if '/zh/products/detail/' in url or '/zh/products/result' in url:
            return True
    except Exception:
        pass
    return False


def _wait_for_cf_challenge(page, timeout=CF_TIMEOUT) -> bool:
    """
    轮询检测 CF 验证页面，一旦检测到就弹窗等真人完成。
    真人确认后继续轮询，直到页面不再显示 CF 且出现 Digikey 正常内容。
    返回 True 表示已通过（或未遇到 CF），False 表示超时。
    """
    global _cf_passed_event
    _cf_passed_event.clear()

    start = time.time()
    popup_shown = False
    user_confirmed = False

    while time.time() - start < timeout:
        is_cf = _is_cf_challenge(page)

        if is_cf:
            if not popup_shown:
                print("  [CF] 检测到 Cloudflare 验证页面，等待真人完成验证...")
                threading.Thread(target=_show_cf_popup, daemon=True).start()
                popup_shown = True

            # 等待真人点击「确定」
            if _cf_passed_event.wait(timeout=CF_CHECK_INTERVAL):
                user_confirmed = True
                print("  [CF] 真人已确认，等待页面跳转...")
                # 不清除 event，让后续不再弹窗
            continue

        # 当前不是 CF 页面
        if popup_shown:
            # 弹过窗：需要确认页面已经正常了才算通过
            if _is_digikey_normal(page):
                print("  [CF] 验证已通过！")
                return True
            # 还没正常，可能是中间过渡页，继续等
            if not user_confirmed:
                # 用户还没确认，CF 就暂时消失了？可能是页面闪烁，继续等
                pass
        else:
            # 没弹过窗也没有 CF，直接通过
            return True

        time.sleep(CF_CHECK_INTERVAL)

    print("  [CF] 等待真人验证超时！")
    return False


# ---------------------------------------------------------------------------
# 主爬虫类
# ---------------------------------------------------------------------------

class DigikeySpiderV2:
    def __init__(self, headless=False):
        """
        Args:
            headless: 是否无头模式。CF 绕过必须用有头模式让真人操作。
        """
        self.headless = headless
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

        self.min_delay = 1.5
        self.max_delay = 3.5

    # ---- 生命周期 ----

    def start(self):
        """启动浏览器"""
        self.playwright = sync_playwright().start()

        # 使用持久化上下文，保留 cookie / localStorage，减少 CF 触发频率
        self.context = self.playwright.chromium.launch_persistent_context(
            user_data_dir='./browser_data',
            headless=self.headless,
            args=[
                '--disable-blink-features=AutomationControlled',
            ],
            no_viewport=True,
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                       '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            locale='zh-CN',
        )

        self.page = self.context.new_page()
        # 隐藏 webdriver 特征
        self.page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => false });
        """)
        print("[*] 浏览器已启动（Playwright）")

    def stop(self):
        """关闭浏览器"""
        if self.context:
            self.context.close()
        if self.playwright:
            self.playwright.stop()
        print("[*] 浏览器已关闭")

    def random_delay(self):
        delay = random.uniform(self.min_delay, self.max_delay)
        time.sleep(delay)

    # ---- 工具 ----

    def load_keywords(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]

    # ---- 核心逻辑 ----

    def _ensure_no_cf(self, timeout=CF_TIMEOUT):
        """阻塞直到 CF 验证通过或超时"""
        return _wait_for_cf_challenge(self.page, timeout=timeout)

    def _goto_with_cf_check(self, url, timeout=30):
        """
        导航到指定 URL，若遇 CF 则等待真人验证后自动重试。
        返回: (final_url, is_success) — is_success=False 表示失败
        """
        try:
            self.page.goto(url, wait_until='domcontentloaded', timeout=timeout * 1000)
        except PlaywrightTimeout:
            print(f"  -> 页面加载超时: {url}")
            return None, False
        except Exception as e:
            print(f"  -> 导航失败: {e}")
            return None, False

        # CF 检测（轻量，几乎无开销）
        if not self._ensure_no_cf():
            print("  -> Cloudflare 验证失败或超时，跳过")
            return None, False

        # 等待关键元素出现（比 networkidle 快得多）
        try:
            self.page.wait_for_selector(
                '[data-testid="in-stock-title"], table tbody tr, a[href*="/detail/"]',
                timeout=timeout * 1000
            )
        except PlaywrightTimeout:
            pass  # 页面可能不含这些元素，忽略

        return self.page.url, True

    def search_product(self, keyword, retry=3):
        """
        搜索产品，等待页面最终稳定后判断结果类型。
        返回: (is_detail_page, is_list_page)
        """
        search_url = f'https://www.digikey.cn/zh/products/result?keywords={quote(keyword)}'

        for attempt in range(retry):
            try:
                final_url, ok = self._goto_with_cf_check(search_url)
                if not ok:
                    continue

                if '/zh/products/detail/' in final_url:
                    print(f"  -> 直接跳转详情页")
                    return True, False

                if '/zh/products/result' in final_url or '/zh/products/category/' in final_url:
                    print(f"  -> 搜索结果列表页")
                    return False, True

                print(f"  -> 未知页面类型: {final_url}")
                return False, False

            except Exception as e:
                if attempt < retry - 1:
                    print(f"  -> 搜索失败，{3*(attempt+1)}秒后重试: {e}")
                    time.sleep(3 * (attempt + 1))
                    continue
                print(f"  -> 搜索 {keyword} 失败: {e}")
                return False, False

    def get_stock_from_page(self):
        """
        从当前页面提取现货数量。
        优先使用 data-testid="in-stock-title"，失败则回退到正则匹配。
        """
        # 方案1：等待 data-testid="in-stock-title" 并直接用 textContent
        try:
            self.page.wait_for_selector('[data-testid="in-stock-title"]', timeout=5000)
            # 用 evaluate 直接读 textContent，比 inner_text() 更稳定
            text = self.page.eval_on_selector(
                '[data-testid="in-stock-title"]',
                'el => el.textContent || ""'
            )
            if text:
                match = re.search(r'([\d,]+)', text)
                if match:
                    return match.group(1).replace(',', '')
        except Exception:
            pass

        # 方案2：通过 Playwright 文本匹配
        try:
            el = self.page.locator('span:has-text("现货")').first
            if el.count() > 0:
                text = el.text_content()
                if text:
                    match = re.search(r'([\d,]+)', text)
                    if match:
                        return match.group(1).replace(',', '')
        except Exception:
            pass

        # 方案3：回退到 BeautifulSoup 正则匹配
        try:
            html = self.page.content()
            soup = BeautifulSoup(html, 'lxml')
            stock_div = soup.select_one('[data-testid="in-stock-title"]')
            if stock_div:
                match = re.search(r'([\d,]+)', stock_div.get_text())
                if match:
                    return match.group(1).replace(',', '')
            # 正则兜底
            patterns = [r'([\d,]+)\s*现货', r'现货\s*[：:]\s*([\d,]+)']
            for span in soup.find_all('span'):
                if span.string and '现货' in span.string:
                    for pat in patterns:
                        match = re.search(pat, span.string)
                        if match:
                            return match.group(1).replace(',', '')
        except Exception:
            pass

        return '未找到'

    def get_detail_from_list(self, keyword, retry=3):
        """
        在搜索结果列表页中查找第一个产品链接并导航到详情页。
        返回 True 表示成功跳转到详情页。
        """
        for attempt in range(retry):
            try:
                selectors = [
                    'a[href*="/zh/products/detail/"]',
                    'a[href*="/detail/"]',
                ]
                clicked = False
                for sel in selectors:
                    link = self.page.locator(sel).first
                    if link.count() > 0:
                        href = link.get_attribute('href')
                        if href:
                            print(f"  -> 点击产品链接跳转详情页")
                            detail_url = href if href.startswith('http') else 'https://www.digikey.cn' + href
                            _, ok = self._goto_with_cf_check(detail_url)
                            if ok and '/zh/products/detail/' in self.page.url:
                                clicked = True
                            break

                if clicked:
                    return True
                print(f"  -> 列表中未找到产品链接")
                return False

            except Exception as e:
                if attempt < retry - 1:
                    print(f"  -> 列表页操作失败，{3*(attempt+1)}秒后重试: {e}")
                    time.sleep(3 * (attempt + 1))
                    continue
                print(f"  -> 列表页操作 {keyword} 失败: {e}")
                return False

        return False

    # ---- 主流程 ----

    def crawl(self, keywords, progress_callback=None):
        results = []

        for i, keyword in enumerate(keywords):
            print(f"\n[{i+1}/{len(keywords)}] 正在爬取: {keyword}")

            try:
                is_detail, is_list = self.search_product(keyword)

                if is_detail:
                    # 已经在详情页，直接提取
                    stock = self.get_stock_from_page()
                elif is_list:
                    # 在列表页，需要点击进入详情页
                    if self.get_detail_from_list(keyword):
                        stock = self.get_stock_from_page()
                    else:
                        stock = '列表页未找到产品'
                else:
                    stock = '未找到产品'
            except Exception as e:
                print(f"  -> 爬取异常: {e}")
                stock = '爬取异常'
                import traceback
                traceback.print_exc()

            print(f"  -> 现货数量: {stock}")

            result = {
                '规格': keyword,
                '现货数量': stock,
            }
            results.append(result)

            if progress_callback:
                progress_callback(i + 1, len(keywords), keyword, result)

            self.random_delay()

        return results

    # ---- 导出 ----

    def export_to_excel(self, data, output_path):
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine='openpyxl')

        wb = load_workbook(output_path)
        ws = wb.active

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        wb.save(output_path)
        print(f"\nExcel 已导出: {output_path}")


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

def main():
    import os
    os.makedirs('output', exist_ok=True)

    spider = DigikeySpiderV2(headless=False)
    spider.start()

    try:
        keywords = spider.load_keywords('config/keywords.txt')
        print(f"已加载 {len(keywords)} 个关键词\n")

        results = spider.crawl(keywords)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'output/digikey_stock_{timestamp}.xlsx'
        spider.export_to_excel(results, output_file)

        print("\n爬取完成！")
    finally:
        spider.stop()


if __name__ == '__main__':
    main()
